"""
System monitoring and analysis functions for master.py
"""

import os
import json
from datetime import datetime
from pathlib import Path

OUTPUT_DIR = Path(r"D:\Generated-Outputs")


# ---------------------------------------------------------------------------
# TASK: List heavy running user processes
# ---------------------------------------------------------------------------

def list_heavy_processes(output_path: Path = None, threshold_cpu: float = 5.0, threshold_memory: float = 50.0):
    """
    Lists all user processes consuming significant CPU/Memory resources.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/heavy_processes.xlsx
        threshold_cpu: Show processes using more than this % CPU
        threshold_memory: Show processes using more than this MB memory
    """
    try:
        import psutil
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("psutil and openpyxl required. Run: pip install psutil openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "heavy_processes.xlsx"

    output_path = Path(output_path)
    print(f"Scanning user processes...")

    processes = []
    for proc in psutil.process_iter(['pid', 'name', 'username', 'cpu_percent', 'memory_info']):
        try:
            pinfo = proc.as_dict(attrs=['pid', 'name', 'username', 'cpu_percent', 'memory_info'])
            cpu = pinfo['cpu_percent'] or 0
            mem_mb = (pinfo['memory_info'].rss / (1024 * 1024)) if pinfo['memory_info'] else 0
            
            if cpu > threshold_cpu or mem_mb > threshold_memory:
                processes.append({
                    'PID': pinfo['pid'],
                    'Process Name': pinfo['name'],
                    'User': pinfo['username'] or 'System',
                    'CPU %': round(cpu, 2),
                    'Memory MB': round(mem_mb, 2),
                })
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass

    # Sort by memory usage
    processes.sort(key=lambda x: x['Memory MB'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Heavy Processes"

    headers = ['PID', 'Process Name', 'User', 'CPU %', 'Memory MB']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Add summary
    ws.insert_rows(1, 2)
    ws['A1'] = f"Heavy Processes Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Threshold: CPU > {threshold_cpu}% or Memory > {threshold_memory}MB | Total: {len(processes)} processes"
    ws['A2'].font = Font(italic=True, size=10)

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    for r_idx, proc in enumerate(processes, start=4):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=proc[key])
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if key in ['PID', 'CPU %', 'Memory MB'] else "left")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    wb.save(output_path)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Found {len(processes)} heavy processes -> {output_path}")


# ---------------------------------------------------------------------------
# TASK: System monitoring graph/report
# ---------------------------------------------------------------------------

def system_monitoring_report(output_path: Path = None, samples: int = 60, interval: int = 1):
    """
    Captures CPU, Memory, Disk, and Network stats over time and creates a report.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/system_monitoring.xlsx
        samples: Number of data samples to collect
        interval: Seconds between samples
    """
    try:
        import psutil
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import time
    except ImportError:
        print("psutil and openpyxl required. Run: pip install psutil openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "system_monitoring.xlsx"

    output_path = Path(output_path)
    print(f"Collecting system monitoring data ({samples} samples)...")

    data = []
    for i in range(samples):
        timestamp = datetime.now().strftime('%H:%M:%S')
        cpu_percent = psutil.cpu_percent(interval=0.1)
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        net = psutil.net_io_counters()
        
        data.append({
            'Timestamp': timestamp,
            'CPU %': cpu_percent,
            'Memory %': memory.percent,
            'Memory Used GB': round(memory.used / (1024**3), 2),
            'Disk %': disk.percent,
            'Disk Used GB': round(disk.used / (1024**3), 2),
            'Net In MB': round(net.bytes_recv / (1024**2), 2),
            'Net Out MB': round(net.bytes_sent / (1024**2), 2),
        })
        
        if i < samples - 1:
            print(f"  [{i+1}/{samples}] CPU: {cpu_percent:.1f}% | Memory: {memory.percent:.1f}% | Disk: {disk.percent:.1f}%")
            time.sleep(interval)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Monitoring"

    headers = list(data[0].keys())
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    ws['A1'] = "System Monitoring Report"
    ws['A1'].font = Font(bold=True, size=14)

    # Calculate averages
    avg_cpu = sum(d['CPU %'] for d in data) / len(data)
    avg_mem = sum(d['Memory %'] for d in data) / len(data)
    
    ws['A1'].value = f"System Monitoring Report - Avg CPU: {avg_cpu:.1f}%, Avg Mem: {avg_mem:.1f}%"

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    for r_idx, row_data in enumerate(data, start=3):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=row_data[key])
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right")

    col_widths = [12, 10, 10, 15, 10, 15, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    wb.save(output_path)
    print(f"System monitoring report -> {output_path}")


# ---------------------------------------------------------------------------
# TASK: List all Chrome extensions
# ---------------------------------------------------------------------------

def list_chrome_extensions(output_path: Path = None):
    """
    Lists all installed Chrome/Chromium extensions.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/chrome_extensions.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl required. Run: pip install openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "chrome_extensions.xlsx"

    output_path = Path(output_path)

    # Chrome extensions directory
    username = os.getenv('USERNAME')
    chrome_path = Path(rf"C:\Users\{username}\AppData\Local\Google\Chrome\User Data\Default\Extensions")
    
    extensions = []
    
    if chrome_path.exists():
        print(f"Scanning Chrome extensions at {chrome_path}...")
        for ext_dir in chrome_path.iterdir():
            if ext_dir.is_dir():
                try:
                    # Get manifest from latest version folder
                    versions = sorted([d for d in ext_dir.iterdir() if d.is_dir()], reverse=True)
                    if versions:
                        manifest_path = versions[0] / "manifest.json"
                        if manifest_path.exists():
                            with open(manifest_path, 'r', encoding='utf-8') as f:
                                manifest = json.load(f)
                                ext_name = manifest.get('name', 'Unknown')
                                ext_version = manifest.get('version', 'Unknown')
                                ext_desc = manifest.get('description', '')
                                permissions = ', '.join(manifest.get('permissions', [])[:5])
                                
                                extensions.append({
                                    'Extension ID': ext_dir.name,
                                    'Name': ext_name,
                                    'Version': ext_version,
                                    'Description': ext_desc[:100],
                                    'Key Permissions': permissions,
                                })
                except:
                    pass
    else:
        print(f"Chrome path not found at {chrome_path}")
        return

    extensions.sort(key=lambda x: x['Name'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chrome Extensions"

    ws['A1'] = f"Chrome Extensions Report - {len(extensions)} extensions"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Extension ID', 'Name', 'Version', 'Description', 'Key Permissions']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    for r_idx, ext in enumerate(extensions, start=3):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=ext[key])
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [32, 30, 10, 50, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    ws.freeze_panes = "A3"
    wb.save(output_path)
    print(f"Found {len(extensions)} Chrome extensions -> {output_path}")


# ---------------------------------------------------------------------------
# TASK: Scan all drives
# ---------------------------------------------------------------------------

def scan_all_drives(output_path: Path = None):
    """
    Scans all drives and reports usage statistics.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/drives_scan.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import psutil
    except ImportError:
        print("openpyxl and psutil required. Run: pip install openpyxl psutil")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "drives_scan.xlsx"

    output_path = Path(output_path)

    drives = []
    print("Scanning all drives...")

    try:
        import psutil
        for partition in psutil.disk_partitions():
            if partition.fstype:  # Skip empty partitions
                try:
                    usage = psutil.disk_usage(partition.mountpoint)
                    used_pct = (usage.used / usage.total * 100) if usage.total > 0 else 0
                    
                    drives.append({
                        'Drive': partition.device,
                        'Mount Point': partition.mountpoint,
                        'File System': partition.fstype,
                        'Total GB': round(usage.total / (1024**3), 2),
                        'Used GB': round(usage.used / (1024**3), 2),
                        'Free GB': round(usage.free / (1024**3), 2),
                        'Used %': round(used_pct, 2),
                    })
                except PermissionError:
                    drives.append({
                        'Drive': partition.device,
                        'Mount Point': partition.mountpoint,
                        'File System': partition.fstype,
                        'Total GB': 'N/A',
                        'Used GB': 'N/A',
                        'Free GB': 'N/A',
                        'Used %': 'Access Denied',
                    })
    except Exception as e:
        print(f"Error scanning drives: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drives Scan"

    ws['A1'] = f"Drive Scan Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Drive', 'Mount Point', 'File System', 'Total GB', 'Used GB', 'Free GB', 'Used %']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    warning_fill = PatternFill("solid", fgColor="FFE699")

    for r_idx, drive in enumerate(drives, start=3):
        used_pct = drive['Used %']
        fill = warning_fill if (isinstance(used_pct, (int, float)) and used_pct > 80) else (alt_fill if r_idx % 2 == 0 else PatternFill())
        
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=drive[key])
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right" if key in ['Total GB', 'Used GB', 'Free GB', 'Used %'] else "left")

    col_widths = [12, 25, 12, 12, 12, 12, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    wb.save(output_path)
    print(f"Drive scan report ({len(drives)} drives) -> {output_path}")


if __name__ == "__main__":
    print("System monitoring functions loaded. Import into master.py or run directly:")
    print("  python system_monitoring.py")
