"""
master.py — padmaimpex1-pixel workspace master script.
Each task is a function. Run a specific function or run all.
Output goes to D:\\Generated-Outputs by default.
"""

import os
import shutil
import sqlite3
import tempfile
import time
import json
from datetime import datetime, timezone
from pathlib import Path
from dateutil import parser as date_parser

OUTPUT_DIR = Path(r"D:\Generated-Outputs")
COPILOT_DB = Path(r"C:\Users\dell\AppData\Roaming\Code\User\globalStorage\github.copilot-chat\session-store.db")

# =========================================================================
# CURRENT SESSION DATETIME - Set at script startup
# =========================================================================
# Current datetime: 2026-07-18T07:27:37.086+05:30 (IST)
SESSION_DATETIME_STR = "2026-07-18T07:27:37.086+05:30"
try:
    SESSION_DATETIME = date_parser.isoparse(SESSION_DATETIME_STR)
except:
    # Fallback to current system time if parsing fails
    SESSION_DATETIME = datetime.now(timezone.utc)

SESSION_TIMESTAMP = SESSION_DATETIME.strftime("%Y%m%d_%H%M%S")
SESSION_DATE = SESSION_DATETIME.strftime("%Y-%m-%d")
SESSION_TIME = SESSION_DATETIME.strftime("%H:%M:%S")
SESSION_TIMEZONE = SESSION_DATETIME.strftime("%z")

print(f"\n{'='*80}")
print(f"MASTER SCRIPT - Session Started")
print(f"  Date:      {SESSION_DATE}")
print(f"  Time:      {SESSION_TIME} {SESSION_TIMEZONE}")
print(f"  Timestamp: {SESSION_TIMESTAMP}")
print(f"  Timezone:  IST (UTC+05:30)")
print(f"{'='*80}\n")


# =========================================================================
# DATETIME UTILITY FUNCTIONS
# =========================================================================

def get_session_datetime():
    """Returns the current session datetime."""
    return SESSION_DATETIME

def get_session_timestamp():
    """Returns the session timestamp in format YYYYMMDD_HHMMSS."""
    return SESSION_TIMESTAMP

def get_session_info():
    """Returns a dictionary with session datetime information."""
    return {
        "datetime": SESSION_DATETIME_STR,
        "date": SESSION_DATE,
        "time": SESSION_TIME,
        "timezone": "IST (UTC+05:30)",
        "timestamp": SESSION_TIMESTAMP,
    }

def log_with_timestamp(message: str, prefix: str = "INFO"):
    """Log a message with the current session timestamp."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] [{prefix}] {message}")

def get_file_timestamp_suffix():
    """Returns a filename-safe timestamp suffix based on session."""
    return SESSION_TIMESTAMP

# Print session info on startup
def print_session_header(title: str = ""):
    """Print a formatted session header with datetime info."""
    header = f"\n{'='*80}"
    if title:
        header += f"\n{title}\n"
    header += f"Session: {SESSION_DATE} at {SESSION_TIME} IST\n"
    header += f"{'='*80}\n"
    print(header)



# =========================================================================
# TASK: Export all Copilot session prompts + responses to Excel
# =========================================================================

def export_copilot_sessions_to_excel(output_path: Path = None, watch: bool = False, interval_seconds: int = 60):
    """
    Reads all Copilot chat sessions and turns from the local session-store SQLite DB
    and writes them to an Excel file with multiple sheets categorized by prompt type.

    Args:
        output_path: Where to save the .xlsx file. Defaults to OUTPUT_DIR/copilot_sessions.xlsx
        watch:       If True, re-exports continuously every `interval_seconds`.
        interval_seconds: How often to re-export when watch=True.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "copilot_sessions.xlsx"

    output_path = Path(output_path)

    print(f"Copilot DB : {COPILOT_DB}")
    print(f"Output XLS : {output_path}")

    if not COPILOT_DB.exists():
        print(f"ERROR: session-store.db not found at {COPILOT_DB}")
        return

    def categorize_prompt(prompt_text):
        """Categorize prompt by keywords."""
        if not prompt_text:
            return "Other"
        p = prompt_text.lower()
        
        # Social Media
        if any(x in p for x in ['twitter', 'facebook', 'instagram', 'tiktok', 'reddit', 'linkedin', 
                                'discord', 'slack', 'whatsapp', 'telegram', 'snapchat', 'pinterest',
                                'youtube', 'twitch', 'mastodon', 'bluesky', 'threads', 'x.com',
                                'social', 'post', 'tweet', 'share', 'viral', 'engagement', 'followers',
                                'hashtag', '@mention', 'dm ', 'dm:', 'channel', 'community', 'group chat']):
            return "Social Media"
        
        # Videos
        if any(x in p for x in ['video', 'mp4', 'avi', 'mov', 'mkv', 'webm', 'flv', 'wmv',
                                'codec', 'ffmpeg', 'premiere', 'davinci', 'resolve', 'final cut',
                                'after effects', 'camtasia', 'obs', 'streaming', 'youtube', 'vimeo',
                                'encoding', 'transcoding', 'resolution', '1080p', '4k', '8k',
                                'frame rate', 'fps', 'bitrate', 'h.264', 'h.265', 'vp9',
                                'video editing', 'motion', 'animation', 'subtitle', 'caption']):
            return "Videos"
        
        # Photos & Images
        if any(x in p for x in ['photo', 'image', 'picture', 'jpg', 'jpeg', 'png', 'gif', 'bmp',
                                'svg', 'webp', 'tiff', 'raw', 'psd', 'ai', 'eps', 'pdf',
                                'photoshop', 'gimp', 'lightroom', 'capture one', 'affinity',
                                'exposure', 'aperture', 'shutter', 'iso', 'white balance',
                                'pixel', 'resolution', 'dpi', 'crop', 'rotate', 'filter',
                                'photography', 'photographic', 'screenshot', 'screenshot', 'canvas',
                                'design', 'graphic', 'visual', 'illustration', 'artwork']):
            return "Photos & Images"
        
        # URLs (check after media categories)
        if any(proto in p for proto in ['http://', 'https://', 'ftp://', 'ssh://', '.com', '.org', '.net', 'github.com', 'gitlab.com']):
            if any(x in p for x in ['http', 'https', 'ftp', '.com', '.org', '.net', 'url', 'link', 'web']):
                return "URLs & Links"
        
        # Ports
        if any(x in p for x in ['port', 'localhost:', ':8080', ':3000', ':5000', ':443', ':80', 'listen', 'binding', 'socket']):
            return "Ports & Networking"
        
        # Backup operations
        if any(x in p for x in ['backup', 'restore', 'archive', 'compress', 'zip', 'tar', 'dump', 'snapshot']):
            return "Backup & Archive"
        
        # Linux/Bash commands (check before generic script detection)
        linux_cmds = [
            'ls ', 'cat ', 'grep ', 'sed ', 'awk ', 'chmod ', 'chown ', 'mkdir ', 'rmdir ',
            'find ', 'locate ', 'which ', 'whereis ', 'apt ', 'yum ', 'pacman ', 'dnf ',
            'systemctl', 'service ', 'sudo ', 'su ', 'ssh ', 'scp ', 'sftp', 'rsync',
            'tar ', 'gzip ', 'bzip2 ', 'unzip ', 'gunzip ', 'tail ', 'head ', 'wc ',
            'cut ', 'paste ', 'sort ', 'uniq ', 'diff ', 'patch ', 'cmp ', 'comm ',
            'file ', 'strings ', 'ldd ', 'nm ', 'readelf ', 'objdump ', 'strace ',
            'ltrace ', 'top ', 'htop ', 'ps ', 'kill ', 'killall ', 'pkill ', 'pgrep ',
            'df ', 'du ', 'mount ', 'umount ', 'fsck ', 'mkfs ', 'fdisk ', 'parted',
            'ifconfig', 'ip addr', 'route ', 'iptables', 'firewall', 'netcat', 'nc ',
            '/bin/', '/usr/bin/', '/etc/', '/var/', '/home/', 'bash', 'sh ', 'zsh'
        ]
        if any(cmd in p for cmd in linux_cmds):
            return "Linux Commands"
        
        # PowerShell commands (check before generic script detection)
        powershell_cmds = [
            'get-childitem', 'get-content', 'select-string', 'invoke-webrequest', 
            'start-process', 'stop-process', 'new-item', 'remove-item', 'copy-item',
            'move-item', 'rename-item', 'set-location', 'test-path', 'get-process',
            'write-host', 'read-host', 'foreach-object', 'where-object', 'group-object',
            'sort-object', 'measure-object', 'get-service', 'start-service', 'stop-service',
            'get-eventlog', 'get-wmiobject', 'add-content', 'clear-content', 'import-module',
            'export-csv', 'import-csv', 'convertto-json', 'convertfrom-json', 'new-psdrive'
        ]
        if any(cmd in p for cmd in powershell_cmds):
            return "PowerShell Commands"
        
        # Windows commands (check before generic script detection)
        windows_cmds = [
            'cmd ', 'dir ', 'del ', 'copy ', 'move ', 'rename ', 'ren ', 'type ', 
            'tasklist', 'taskkill', 'systeminfo', 'ipconfig', 'ping ', 'netstat', 
            'schtasks', 'wmic', 'reg ', 'bitsadmin', 'findstr', 'fc ', 'for /f',
            'pushd', 'popd', 'cd ', 'mkdir', 'rmdir', 'robocopy', 'xcopy', 'comp ',
            'diskpart', 'format ', 'chkdsk', 'defrag', 'cipher', 'takeown', 'icacls',
            'attrib ', 'date /t', 'time /t', 'nslookup', 'tracert', 'route ', 'arp '
        ]
        if any(cmd in p for cmd in windows_cmds):
            return "Windows Commands"
        
        # Repository-specific work (check for common repo patterns)
        repo_patterns = [
            'starter-repo', 'windows-disk-cleanup', 'pixel', 'padmaimpex', 'myapp',
            'vipingit', 'misc', 'd:\\gitrepos', 'github.com/', 'gitlab.com/', 'bitbucket.org/',
            'src/', 'test/', 'docs/', 'readme.md', '.github/', '.gitignore'
        ]
        if any(repo in p for repo in repo_patterns):
            return "Repository Work"
        
        if any(x in p for x in ['move', 'copy', 'delete', 'rename', 'organize', 'folder', 'directory']):
            return "File Operations"
        if any(x in p for x in ['python', 'script', 'code', 'function', 'class', 'def']):
            return "Code & Scripts"
        if any(x in p for x in ['git', 'github', 'commit', 'push', 'merge', 'branch']):
            return "Git & GitHub"
        if any(x in p for x in ['install', 'pip', 'npm', 'package', 'dependency', 'requirements']):
            return "Package Management"
        if any(x in p for x in ['fix', 'bug', 'error', 'debug', 'issue', 'problem']):
            return "Debugging & Fixes"
        if any(x in p for x in ['create', 'generate', 'write', 'build', 'implement']):
            return "Creation & Generation"
        if any(x in p for x in ['read', 'view', 'check', 'list', 'show', 'display', 'inspect']):
            return "Inspection & Analysis"
        if any(x in p for x in ['excel', 'xls', 'csv', 'export', 'data', 'database']):
            return "Data & Export"
        if any(x in p for x in ['instruction', 'config', 'settings', 'default', 'convention']):
            return "Configuration & Conventions"
        return "Other"

    def do_export():
        # Copy DB to temp location to avoid locking the live file
        tmp = Path(tempfile.gettempdir()) / "copilot-session-store-copy.db"
        shutil.copy2(COPILOT_DB, tmp)

        con = sqlite3.connect(tmp)
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        rows = cur.execute("""
            SELECT
                s.id            AS session_id,
                s.cwd           AS cwd,
                s.repository    AS repository,
                s.branch        AS branch,
                s.summary       AS session_summary,
                s.created_at    AS session_created,
                s.updated_at    AS session_updated,
                t.turn_index    AS turn_index,
                t.user_message  AS prompt,
                t.assistant_response AS response,
                t.timestamp     AS turn_timestamp
            FROM sessions s
            JOIN turns t ON t.session_id = s.id
            ORDER BY t.timestamp DESC, s.created_at DESC, t.turn_index DESC
        """).fetchall()

        con.close()
        tmp.unlink(missing_ok=True)

        # Group by category
        categorized = {}
        for row in rows:
            cat = categorize_prompt(row['prompt'])
            if cat not in categorized:
                categorized[cat] = []
            categorized[cat].append(row)

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        headers = [
            "Session ID", "CWD", "Repository", "Branch",
            "Session Summary", "Session Created", "Session Updated",
            "Turn #", "Prompt (User)", "Response (Copilot)", "Turn Timestamp"
        ]

        # Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary['A1'] = "Copilot Sessions Report"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A3'] = "Category"
        ws_summary['B3'] = "Turn Count"
        ws_summary['A3'].font = Font(bold=True)
        ws_summary['B3'].font = Font(bold=True)
        row_num = 4
        for cat in sorted(categorized.keys()):
            ws_summary[f'A{row_num}'] = cat
            ws_summary[f'B{row_num}'] = len(categorized[cat])
            row_num += 1
        ws_summary['A2'] = f"Total turns: {len(rows)} | Total categories: {len(categorized)}"
        
        # Keywords sheet
        ws_keywords = wb.create_sheet("Keywords", 1)
        ws_keywords['A1'] = "Keyword Analysis"
        ws_keywords['A1'].font = Font(bold=True, size=14)
        
        keyword_dict = {}
        keyword_patterns = [
            'python', 'javascript', 'typescript', 'java', 'csharp', 'rust', 'go', 'rust',
            'docker', 'kubernetes', 'aws', 'azure', 'gcp', 'cloud', 'database', 'sql', 'nosql',
            'api', 'rest', 'graphql', 'websocket', 'microservice', 'serverless', 'lambda',
            'testing', 'unit test', 'integration test', 'e2e', 'jest', 'pytest', 'mocha',
            'ci/cd', 'jenkins', 'gitlab-ci', 'github actions', 'devops', 'terraform',
            'monitoring', 'logging', 'prometheus', 'grafana', 'elk', 'datadog', 'newrelic',
            'security', 'encryption', 'ssl', 'tls', 'oauth', 'jwt', 'authentication',
            'performance', 'optimization', 'cache', 'redis', 'memcached', 'caching',
            'frontend', 'react', 'vue', 'angular', 'svelte', 'html', 'css', 'webpack',
            'backend', 'nodejs', 'django', 'flask', 'fastapi', 'spring', 'rails',
            'database', 'postgresql', 'mysql', 'mongodb', 'dynamodb', 'cassandra',
            'version control', 'git', 'github', 'gitlab', 'bitbucket', 'svn',
            'linux', 'ubuntu', 'debian', 'centos', 'rhel', 'alpine', 'docker',
            'windows', 'powershell', 'batch', 'cmd', 'wsl', 'hyper-v',
            'mac', 'macos', 'osx', 'homebrew', 'xcode', 'swift', 'objective-c'
        ]
        
        for row in rows:
            prompt = (row['prompt'] or '').lower()
            response = (row['response'] or '').lower()
            combined = prompt + ' ' + response
            for kw in keyword_patterns:
                if kw in combined:
                    if kw not in keyword_dict:
                        keyword_dict[kw] = {'count': 0, 'sessions': set(), 'categories': []}
                    keyword_dict[kw]['count'] += 1
                    keyword_dict[kw]['sessions'].add(row['session_id'])
                    cat = categorize_prompt(row['prompt'])
                    if cat not in keyword_dict[kw]['categories']:
                        keyword_dict[kw]['categories'].append(cat)
        
        # Sort keywords by frequency
        sorted_keywords = sorted(keyword_dict.items(), key=lambda x: x[1]['count'], reverse=True)
        
        ws_keywords['A3'] = "Keyword"
        ws_keywords['B3'] = "Count"
        ws_keywords['C3'] = "Sessions"
        ws_keywords['D3'] = "Categories"
        for col in ['A', 'B', 'C', 'D']:
            ws_keywords[f'{col}3'].font = Font(bold=True)
            ws_keywords[f'{col}3'].fill = PatternFill("solid", fgColor="1F4E79")
            ws_keywords[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        
        row_num = 4
        for kw, data in sorted_keywords:
            ws_keywords[f'A{row_num}'] = kw
            ws_keywords[f'B{row_num}'] = data['count']
            ws_keywords[f'C{row_num}'] = len(data['sessions'])
            ws_keywords[f'D{row_num}'] = ', '.join(data['categories'][:3])
            row_num += 1
        
        ws_keywords.column_dimensions['A'].width = 20
        ws_keywords.column_dimensions['B'].width = 10
        ws_keywords.column_dimensions['C'].width = 12
        ws_keywords.column_dimensions['D'].width = 50

        # One sheet per category
        for cat in sorted(categorized.keys()):
            cat_rows = categorized[cat]
            sheet_name = cat[:31]  # Excel sheet name limit
            ws = wb.create_sheet(sheet_name)

            # Header row styling
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")

            # Data rows
            alt_fill = PatternFill("solid", fgColor="D9E1F2")
            for r_idx, row in enumerate(cat_rows, start=2):
                fill = alt_fill if r_idx % 2 == 0 else PatternFill()
                values = [
                    row["session_id"], row["cwd"], row["repository"], row["branch"],
                    row["session_summary"], row["session_created"], row["session_updated"],
                    row["turn_index"], row["prompt"], row["response"], row["turn_timestamp"]
                ]
                for c_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val or "")
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Column widths
            col_widths = [36, 30, 40, 12, 40, 22, 22, 8, 60, 80, 22]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(w, 80)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{now}] Exported {len(rows)} turns across {len(set(r['session_id'] for r in rows))} sessions, {len(categorized)} categories -> {output_path}")

    if watch:
        print(f"Watch mode ON -- refreshing every {interval_seconds}s. Press Ctrl+C to stop.")
        while True:
            try:
                do_export()
                time.sleep(interval_seconds)
            except KeyboardInterrupt:
                print("Stopped.")
                break
    else:
        do_export()


# ---------------------------------------------------------------------------
# TASK: Sync vipingit1 repos to padmaimpex1-pixel on GitHub
# ---------------------------------------------------------------------------

def sync_gitrepos_to_padmaimpex(root: str = r"D:\GitRepos", target_owner: str = "padmaimpex1-pixel", dry_run: bool = False):
    """
    Scans all Git repos under `root`, and for any repo whose remote origin
    points to `vipingit1`, re-points it to `padmaimpex1-pixel` and pushes.

    Skips:
      - Duplicate folders containing '-from-' in their name
      - Third-party repos (not vipingit1 or padmaimpex1-pixel)
      - Repos already on padmaimpex1-pixel (just pushes them)

    Args:
        root:         Root folder to scan for git repos.
        target_owner: GitHub account to push to.
        dry_run:      If True, print what would happen without making changes.
    """
    import subprocess, json, urllib.request, urllib.error

    def run(cmd, cwd=None, capture=True):
        r = subprocess.run(cmd, cwd=cwd, capture_output=capture, text=True)
        return r.stdout.strip(), r.stderr.strip(), r.returncode

    def get_git_token():
        out, _, _ = run(['git', 'credential', 'fill'],)
        # Try to get token from Windows credential manager via git
        # Use a known working URL to extract token
        import subprocess as sp
        proc = sp.Popen(['git', 'credential', 'fill'], stdin=sp.PIPE, stdout=sp.PIPE, stderr=sp.PIPE, text=True)
        stdout, _ = proc.communicate(input="protocol=https\nhost=github.com\n\n")
        for line in stdout.splitlines():
            if line.startswith('password='):
                return line[len('password='):]
        return None

    def repo_exists_on_github(owner, repo_name, token):
        url = f"https://api.github.com/repos/{owner}/{repo_name}"
        req = urllib.request.Request(url, headers={"Authorization": f"token {token}", "User-Agent": "master.py"})
        try:
            urllib.request.urlopen(req, timeout=10)
            return True
        except urllib.error.HTTPError as e:
            return e.code != 404

    def create_repo_on_github(owner, repo_name, token, private=True):
        url = "https://api.github.com/user/repos"
        data = json.dumps({"name": repo_name, "private": private, "auto_init": False}).encode()
        req = urllib.request.Request(url, data=data, method="POST",
                                     headers={"Authorization": f"token {token}",
                                              "Content-Type": "application/json",
                                              "User-Agent": "master.py"})
        try:
            urllib.request.urlopen(req, timeout=15)
            return True, "created"
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            if "already exists" in body or e.code == 422:
                return True, "already_exists"
            return False, f"HTTP {e.code}: {body[:200]}"

    # Find all git repos
    root_path = Path(root)
    git_dirs = [p.parent for p in root_path.rglob('.git') if p.is_dir()]
    git_dirs = sorted(set(git_dirs))

    token = get_git_token()
    if not token:
        print("ERROR: Could not retrieve GitHub token from git credential manager.")
        return

    results = []
    skipped = []

    for repo_path in git_dirs:
        name = repo_path.name

        # Skip duplicates
        if '-from-' in name:
            skipped.append((str(repo_path), "duplicate"))
            continue

        remote_out, _, rc = run(['git', 'remote', 'get-url', 'origin'], cwd=str(repo_path))
        if rc != 0 or not remote_out:
            skipped.append((str(repo_path), "no_remote"))
            continue

        remote = remote_out.strip().rstrip('.git')
        remote_lower = remote.lower()

        # Already on target
        if f"github.com/{target_owner.lower()}/" in remote_lower:
            repo_name = remote.split('/')[-1]
            status = "already_target"
        elif "github.com/vipingit1/" in remote_lower:
            repo_name = remote.split('/')[-1]
            status = "vipingit1"
        else:
            skipped.append((str(repo_path), f"third_party: {remote}"))
            continue

        branch_out, _, _ = run(['git', 'branch', '--show-current'], cwd=str(repo_path))
        branch = branch_out.strip() or "main"
        new_remote = f"https://github.com/{target_owner}/{repo_name}.git"

        print(f"\n[{name}]")
        print(f"  Repo   : {repo_name}")
        print(f"  Remote : {remote} -> {new_remote}")
        print(f"  Branch : {branch}")

        if dry_run:
            print("  DRY_RUN: skipping actual changes")
            results.append({"repo": name, "action": "dry_run", "target": new_remote})
            continue

        # Ensure repo exists on padmaimpex1-pixel
        exists = repo_exists_on_github(target_owner, repo_name, token)
        if not exists:
            ok, msg = create_repo_on_github(target_owner, repo_name, token, private=True)
            print(f"  Create : {msg}")
            if not ok:
                results.append({"repo": name, "action": "create_failed", "error": msg})
                continue
        else:
            print(f"  Create : repo already exists on GitHub")

        # Update remote
        if status == "vipingit1":
            run(['git', 'remote', 'set-url', 'origin', new_remote], cwd=str(repo_path))
            print(f"  Remote updated to {new_remote}")

        # Push
        push_out, push_err, push_rc = run(['git', 'push', 'origin', branch, '--force-with-lease'], cwd=str(repo_path))
        if push_rc == 0:
            print(f"  PUSHED OK")
            results.append({"repo": name, "action": "pushed", "target": new_remote})
        else:
            # Try without --force-with-lease
            push_out2, push_err2, push_rc2 = run(['git', 'push', '--set-upstream', 'origin', branch], cwd=str(repo_path))
            if push_rc2 == 0:
                print(f"  PUSHED OK (set-upstream)")
                results.append({"repo": name, "action": "pushed", "target": new_remote})
            else:
                print(f"  PUSH FAILED: {push_err or push_err2}")
                results.append({"repo": name, "action": "push_failed", "error": push_err or push_err2})

    print(f"\n\n=== SYNC SUMMARY ===")
    pushed    = [r for r in results if r.get('action') == 'pushed']
    failed    = [r for r in results if 'failed' in r.get('action','')]
    print(f"  Pushed  : {len(pushed)}")
    print(f"  Failed  : {len(failed)}")
    print(f"  Skipped : {len(skipped)}")
    if failed:
        print("\n  FAILURES:")
        for f in failed:
            print(f"    {f['repo']}: {f.get('error','')[:120]}")
    if skipped:
        print("\n  SKIPPED:")
        for s, reason in skipped:
            print(f"    {Path(s).name}: {reason}")


# ---------------------------------------------------------------------------
# TASK: Build Home-sheet hyperlinks for all workbook tabs
# ---------------------------------------------------------------------------

def build_home_sheet_hyperlinks(
    workbook_path: Path = Path(r"D:\documents\master_working.xlsm"),
    home_sheet_name: str = "Home",
    output_path: Path = None,
    columns: int = 0,
    start_row: int = 2,
    start_col: int = 1,
    clear_right_preview: bool = True,
    include_home_variants: bool = False,
    include_hidden_sheets: bool = False,
    create_backup: bool = True,
    target_visible_rows: int = 45,
    max_visible_cols: int = 18,
    zoom_scale: int = 55,
):
    """
    Writes clickable tab hyperlinks on the Home sheet for all workbook sheets.

    The links are written as a spread grid starting from row 2, sized to fit
    within a single-page view as much as possible.
    Optionally clears the right-preview panel area (default: columns D:H).
    """
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return

    workbook_path = Path(workbook_path)
    if output_path is None:
        output_path = workbook_path
    else:
        output_path = Path(output_path)

    if not workbook_path.exists():
        print(f"ERROR: workbook not found: {workbook_path}")
        return

    wb = openpyxl.load_workbook(workbook_path, keep_vba=True)
    if home_sheet_name not in wb.sheetnames:
        print(f"ERROR: home sheet not found: {home_sheet_name}")
        return

    ws = wb[home_sheet_name]

    # Build sheet list in workbook order.
    sheet_names = []
    for s in wb.worksheets:
        name = s.title
        if name == home_sheet_name:
            continue
        if not include_home_variants and name.lower().startswith("home"):
            continue
        if not include_hidden_sheets and s.sheet_state != "visible":
            continue
        sheet_names.append(name)

    if not sheet_names:
        print("No target sheets found to link.")
        return

    # Auto-calculate spread columns so links fit in a single page view.
    if columns <= 0:
        columns = max(1, (len(sheet_names) + target_visible_rows - 1) // target_visible_rows)
        columns = min(columns, max_visible_cols)

    # Clear existing home link grid area.
    rows_needed = (len(sheet_names) + columns - 1) // columns
    end_row = max(ws.max_row, start_row + max(rows_needed, target_visible_rows) + 5)
    end_col = start_col + max(columns, max_visible_cols) - 1
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.hyperlink = None

    # Write formulas as hyperlinks to each sheet tab.
    for idx, sheet_name in enumerate(sheet_names):
        r = start_row + (idx // columns)
        c = start_col + (idx % columns)
        sheet_escaped = sheet_name.replace("'", "''")
        text_escaped = sheet_name.replace('"', '""')
        ws.cell(r, c).value = f'=HYPERLINK("#\'{sheet_escaped}\'!A1","{text_escaped}")'

    # Format the visible link area to improve one-page readability.
    for c in range(start_col, start_col + columns):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12
    for r in range(start_row, start_row + rows_needed + 1):
        ws.row_dimensions[r].height = 18

    # Make sheet view and print setup fit in a single page as much as possible.
    ws.sheet_view.zoomScale = zoom_scale
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Remove right-side preview content if requested.
    if clear_right_preview:
        preview_start_col = 4  # D
        preview_end_col = 8    # H
        preview_end_row = max(ws.max_row, start_row + rows_needed + 20)
        for r in range(1, preview_end_row + 1):
            for c in range(preview_start_col, preview_end_col + 1):
                ws.cell(r, c).value = None
                ws.cell(r, c).hyperlink = None

    if create_backup and output_path == workbook_path:
        backup_path = workbook_path.with_name(
            f"{workbook_path.stem}.backup-{datetime.now().strftime('%Y%m%d_%H%M%S')}{workbook_path.suffix}"
        )
        shutil.copy2(workbook_path, backup_path)
        print(f"Backup created: {backup_path}")

    wb.save(output_path)
    print(f"Updated Home hyperlinks for {len(sheet_names)} tabs -> {output_path}")
    print(
        f"Grid: start row {start_row}, columns {columns}, rows {rows_needed}, "
        f"zoom {zoom_scale}%, right preview cleared: {clear_right_preview}"
    )


# ---------------------------------------------------------------------------
# MENU
# ---------------------------------------------------------------------------

TASKS = {
    "1": ("Export Copilot sessions to Excel (once)", lambda: export_copilot_sessions_to_excel()),
    "2": ("Export Copilot sessions to Excel (watch/continuous)", lambda: export_copilot_sessions_to_excel(watch=True, interval_seconds=60)),
    "3": ("Sync vipingit1 repos to padmaimpex1-pixel (dry run)", lambda: sync_gitrepos_to_padmaimpex(dry_run=True)),
    "4": ("Sync vipingit1 repos to padmaimpex1-pixel (LIVE)", lambda: sync_gitrepos_to_padmaimpex(dry_run=False)),
    "5": ("Build Home-sheet hyperlinks for all tabs (clear right preview)", lambda: build_home_sheet_hyperlinks()),
}

if __name__ == "__main__":
    print("\n=== padmaimpex1-pixel master.py ===\n")
    for key, (label, _) in TASKS.items():
        print(f"  [{key}] {label}")
    print()
    choice = input("Select task (or press Enter for task 1): ").strip() or "1"
    if choice in TASKS:
        TASKS[choice][1]()
    else:
        print(f"Unknown choice: {choice}")


# ---------------------------------------------------------------------------
# TASK: List heavy running user processes
# ---------------------------------------------------------------------------

def list_heavy_processes(output_path: Path = None, threshold_cpu: float = 5.0, threshold_memory: float = 50.0):
    """
    Lists all user processes consuming significant CPU/Memory resources.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/heavy_processes.xlsx
        threshold_cpu: Show processes using more than this % CPU
        threshold_memory: Show processes using more than this MB memory
    """
    try:
        import psutil
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("psutil and openpyxl required. Run: pip install psutil openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "heavy_processes.xlsx"

    output_path = Path(output_path)
    print(f"Scanning user processes...")

    processes = []
    for proc in psutil.process_iter(['pid', 'name', 'username', 'cpu_percent', 'memory_info']):
        try:
            pinfo = proc.as_dict(attrs=['pid', 'name', 'username', 'cpu_percent', 'memory_info'])
            cpu = pinfo['cpu_percent'] or 0
            mem_mb = (pinfo['memory_info'].rss / (1024 * 1024)) if pinfo['memory_info'] else 0
            
            if cpu > threshold_cpu or mem_mb > threshold_memory:
                processes.append({
                    'PID': pinfo['pid'],
                    'Process Name': pinfo['name'],
                    'User': pinfo['username'] or 'System',
                    'CPU %': round(cpu, 2),
                    'Memory MB': round(mem_mb, 2),
                })
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass

    # Sort by memory usage
    processes.sort(key=lambda x: x['Memory MB'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Heavy Processes"

    headers = ['PID', 'Process Name', 'User', 'CPU %', 'Memory MB']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Add summary
    ws.insert_rows(1, 2)
    ws['A1'] = f"Heavy Processes Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Threshold: CPU > {threshold_cpu}% or Memory > {threshold_memory}MB | Total: {len(processes)} processes"
    ws['A2'].font = Font(italic=True, size=10)

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    for r_idx, proc in enumerate(processes, start=4):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=proc[key])
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if key in ['PID', 'CPU %', 'Memory MB'] else "left")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    wb.save(output_path)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Found {len(processes)} heavy processes -> {output_path}")


# ---------------------------------------------------------------------------
# TASK: System monitoring graph/report
# ---------------------------------------------------------------------------

def system_monitoring_report(output_path: Path = None, samples: int = 60, interval: int = 1):
    """
    Captures CPU, Memory, Disk, and Network stats over time and creates a report.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/system_monitoring.xlsx
        samples: Number of data samples to collect
        interval: Seconds between samples
    """
    try:
        import psutil
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("psutil and openpyxl required. Run: pip install psutil openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "system_monitoring.xlsx"

    output_path = Path(output_path)
    print(f"Collecting system monitoring data ({samples} samples)...")

    data = []
    for i in range(samples):
        timestamp = datetime.now().strftime('%H:%M:%S')
        cpu_percent = psutil.cpu_percent(interval=0.1)
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        net = psutil.net_io_counters()
        
        data.append({
            'Timestamp': timestamp,
            'CPU %': cpu_percent,
            'Memory %': memory.percent,
            'Memory Used GB': round(memory.used / (1024**3), 2),
            'Disk %': disk.percent,
            'Disk Used GB': round(disk.used / (1024**3), 2),
            'Net In MB': round(net.bytes_recv / (1024**2), 2),
            'Net Out MB': round(net.bytes_sent / (1024**2), 2),
        })
        
        if i < samples - 1:
            print(f"  [{i+1}/{samples}] CPU: {cpu_percent:.1f}% | Memory: {memory.percent:.1f}% | Disk: {disk.percent:.1f}%")
            time.sleep(interval)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Monitoring"

    headers = list(data[0].keys())
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    ws['A1'] = "System Monitoring Report"
    ws['A1'].font = Font(bold=True, size=14)

    # Calculate averages
    avg_cpu = sum(d['CPU %'] for d in data) / len(data)
    avg_mem = sum(d['Memory %'] for d in data) / len(data)
    
    ws['A1'].value = f"System Monitoring Report - Avg CPU: {avg_cpu:.1f}%, Avg Mem: {avg_mem:.1f}%"

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    for r_idx, row_data in enumerate(data, start=3):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=row_data[key])
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right")

    col_widths = [12, 10, 10, 15, 10, 15, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    wb.save(output_path)
    print(f"System monitoring report -> {output_path}")


# ---------------------------------------------------------------------------
# TASK: List all Chrome extensions
# ---------------------------------------------------------------------------

def list_chrome_extensions(output_path: Path = None):
    """
    Lists all installed Chrome/Chromium extensions.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/chrome_extensions.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl required. Run: pip install openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "chrome_extensions.xlsx"

    output_path = Path(output_path)

    # Chrome extensions directory
    username = os.getenv('USERNAME')
    chrome_path = Path(rf"C:\Users\{username}\AppData\Local\Google\Chrome\User Data\Default\Extensions")
    
    extensions = []
    
    if chrome_path.exists():
        print(f"Scanning Chrome extensions at {chrome_path}...")
        for ext_dir in chrome_path.iterdir():
            if ext_dir.is_dir():
                try:
                    # Get manifest from latest version folder
                    versions = sorted([d for d in ext_dir.iterdir() if d.is_dir()], reverse=True)
                    if versions:
                        manifest_path = versions[0] / "manifest.json"
                        if manifest_path.exists():
                            with open(manifest_path, 'r', encoding='utf-8') as f:
                                manifest = json.load(f)
                                ext_name = manifest.get('name', 'Unknown')
                                ext_version = manifest.get('version', 'Unknown')
                                ext_desc = manifest.get('description', '')
                                permissions = ', '.join(manifest.get('permissions', [])[:5])
                                
                                extensions.append({
                                    'Extension ID': ext_dir.name,
                                    'Name': ext_name,
                                    'Version': ext_version,
                                    'Description': ext_desc[:100],
                                    'Key Permissions': permissions,
                                })
                except:
                    pass
    else:
        print(f"Chrome path not found at {chrome_path}")
        return

    extensions.sort(key=lambda x: x['Name'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chrome Extensions"

    ws['A1'] = f"Chrome Extensions Report - {len(extensions)} extensions"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Extension ID', 'Name', 'Version', 'Description', 'Key Permissions']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    for r_idx, ext in enumerate(extensions, start=3):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=ext[key])
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [32, 30, 10, 50, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    ws.freeze_panes = "A3"
    wb.save(output_path)
    print(f"Found {len(extensions)} Chrome extensions -> {output_path}")


# ---------------------------------------------------------------------------
# TASK: Scan all drives
# ---------------------------------------------------------------------------

def scan_all_drives(output_path: Path = None):
    """
    Scans all drives and reports usage statistics.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/drives_scan.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import psutil
    except ImportError:
        print("openpyxl and psutil required. Run: pip install openpyxl psutil")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "drives_scan.xlsx"

    output_path = Path(output_path)

    drives = []
    print("Scanning all drives...")

    try:
        import psutil
        for partition in psutil.disk_partitions():
            if partition.fstype:  # Skip empty partitions
                try:
                    usage = psutil.disk_usage(partition.mountpoint)
                    used_pct = (usage.used / usage.total * 100) if usage.total > 0 else 0
                    
                    drives.append({
                        'Drive': partition.device,
                        'Mount Point': partition.mountpoint,
                        'File System': partition.fstype,
                        'Total GB': round(usage.total / (1024**3), 2),
                        'Used GB': round(usage.used / (1024**3), 2),
                        'Free GB': round(usage.free / (1024**3), 2),
                        'Used %': round(used_pct, 2),
                    })
                except PermissionError:
                    drives.append({
                        'Drive': partition.device,
                        'Mount Point': partition.mountpoint,
                        'File System': partition.fstype,
                        'Total GB': 'N/A',
                        'Used GB': 'N/A',
                        'Free GB': 'N/A',
                        'Used %': 'Access Denied',
                    })
    except Exception as e:
        print(f"Error scanning drives: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drives Scan"

    ws['A1'] = f"Drive Scan Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Drive', 'Mount Point', 'File System', 'Total GB', 'Used GB', 'Free GB', 'Used %']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    warning_fill = PatternFill("solid", fgColor="FFE699")

    for r_idx, drive in enumerate(drives, start=3):
        used_pct = drive['Used %']
        fill = warning_fill if (isinstance(used_pct, (int, float)) and used_pct > 80) else (alt_fill if r_idx % 2 == 0 else PatternFill())
        
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col, value=drive[key])
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right" if key in ['Total GB', 'Used GB', 'Free GB', 'Used %'] else "left")

    col_widths = [12, 25, 12, 12, 12, 12, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    wb.save(output_path)
    print(f"Drive scan report ({len(drives)} drives) -> {output_path}")


# ---------------------------------------------------------------------------
# TASK: Deep disk usage analysis - where is all the space going?
# ---------------------------------------------------------------------------

def analyze_disk_usage_by_directory(target_drive: str = "C:", output_path: Path = None, top_dirs: int = 50):
    """
    Analyzes disk usage by directory to find where all the space is going.
    
    Args:
        target_drive: Drive to analyze (default: C:)
        output_path: Where to save the report
        top_dirs: Number of top directories to show (default: 50)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("openpyxl required. Run: pip install openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / f"disk_usage_{target_drive.replace(':', '')}_drive.xlsx"

    output_path = Path(output_path)

    print(f"Analyzing disk usage on {target_drive} drive...")
    print("This may take several minutes...")

    dir_sizes = {}
    errors = 0
    
    # Common directories to skip
    skip_dirs = [
        'System Volume Information',
        '$Recycle.Bin',
        'hiberfil.sys',
        'pagefile.sys',
        'swapfile.sys'
    ]

    try:
        root_path = Path(target_drive + "\\")
        
        for item in root_path.iterdir():
            if item.is_dir():
                dir_name = item.name
                
                if dir_name in skip_dirs:
                    continue
                
                try:
                    total_size = 0
                    file_count = 0
                    
                    for root, dirs, files in os.walk(item):
                        # Skip problematic directories
                        dirs[:] = [d for d in dirs if d not in skip_dirs]
                        
                        for file in files:
                            try:
                                filepath = Path(root) / file
                                total_size += filepath.stat().st_size
                                file_count += 1
                            except (PermissionError, OSError):
                                errors += 1
                    
                    if total_size > 0:
                        dir_sizes[dir_name] = {
                            'Size GB': total_size / (1024**3),
                            'Size MB': total_size / (1024**2),
                            'Files': file_count,
                        }
                        
                        if len(dir_sizes) % 5 == 0:
                            print(f"  Scanned {len(dir_sizes)} directories...")
                
                except PermissionError:
                    errors += 1
            elif item.is_file():
                # Files at root
                try:
                    size = item.stat().st_size
                    if size > 0:
                        if '[Root Files]' not in dir_sizes:
                            dir_sizes['[Root Files]'] = {'Size GB': 0, 'Size MB': 0, 'Files': 0}
                        dir_sizes['[Root Files]']['Size GB'] += size / (1024**3)
                        dir_sizes['[Root Files]']['Size MB'] += size / (1024**2)
                        dir_sizes['[Root Files]']['Files'] += 1
                except (PermissionError, OSError):
                    errors += 1

    except Exception as e:
        print(f"Error during analysis: {e}")

    # Sort by size
    sorted_dirs = sorted(dir_sizes.items(), key=lambda x: x[1]['Size GB'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Disk Usage"

    total_gb = sum(d[1]['Size GB'] for d in sorted_dirs)
    total_files = sum(d[1]['Files'] for d in sorted_dirs)

    ws['A1'] = f"{target_drive} Drive Disk Usage Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Total: {total_gb:.2f} GB across {total_files:,} files"
    ws['A2'].font = Font(italic=True, size=10)

    headers = ['Directory', 'Size GB', 'Size MB', 'Percent %', 'File Count', 'Avg File Size MB']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Color scheme based on usage
    high_fill = PatternFill("solid", fgColor="FF6B6B")  # Red >20%
    med_fill = PatternFill("solid", fgColor="FFE699")   # Yellow 10-20%
    low_fill = PatternFill("solid", fgColor="D9E1F2")   # Light blue

    for r_idx, (dir_name, data) in enumerate(sorted_dirs[:top_dirs], start=4):
        pct = (data['Size GB'] / total_gb * 100) if total_gb > 0 else 0
        avg_file_size = data['Size MB'] / data['Files'] if data['Files'] > 0 else 0

        # Choose color based on percentage
        if pct > 20:
            fill = high_fill
        elif pct > 10:
            fill = med_fill
        else:
            fill = low_fill

        ws.cell(row=r_idx, column=1, value=dir_name).fill = fill
        ws.cell(row=r_idx, column=1).font = Font(bold=True if pct > 15 else False)
        ws.cell(row=r_idx, column=2, value=round(data['Size GB'], 2)).fill = fill
        ws.cell(row=r_idx, column=3, value=round(data['Size MB'], 0)).fill = fill
        ws.cell(row=r_idx, column=4, value=round(pct, 1)).fill = fill
        ws.cell(row=r_idx, column=5, value=data['Files']).fill = fill
        ws.cell(row=r_idx, column=6, value=round(avg_file_size, 2)).fill = fill

        for col in range(2, 7):
            ws.cell(row=r_idx, column=col).alignment = Alignment(horizontal="right")

    col_widths = [35, 12, 12, 12, 12, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = f"{target_drive} Drive Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)

    summary_items = [
        ('Total Capacity', f"{total_gb:.2f} GB"),
        ('Total Used', f"{total_gb:.2f} GB"),
        ('Total Files', f"{total_files:,}"),
        ('Directories Scanned', len(dir_sizes)),
        ('Access Errors', errors),
        ('Average File Size', f"{(total_gb * 1024 / total_files):.2f} MB" if total_files > 0 else "N/A"),
    ]

    ws_summary['A3'] = 'Metric'
    ws_summary['B3'] = 'Value'
    for col in ['A', 'B']:
        ws_summary[f'{col}3'].font = Font(bold=True)
        ws_summary[f'{col}3'].fill = PatternFill("solid", fgColor="1F4E79")
        ws_summary[f'{col}3'].font = Font(bold=True, color="FFFFFF")

    row_num = 4
    for metric, value in summary_items:
        ws_summary[f'A{row_num}'] = metric
        ws_summary[f'B{row_num}'] = value
        row_num += 1

    # Top categories
    ws_summary['A12'] = "Top 15 Directories by Size"
    ws_summary['A12'].font = Font(bold=True, size=12)

    ws_summary['A14'] = "Directory"
    ws_summary['B14'] = "Size GB"
    ws_summary['C14'] = "% of Total"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}14'].font = Font(bold=True)
        ws_summary[f'{col}14'].fill = PatternFill("solid", fgColor="1F4E79")
        ws_summary[f'{col}14'].font = Font(bold=True, color="FFFFFF")

    row_num = 15
    for dir_name, data in sorted_dirs[:15]:
        pct = (data['Size GB'] / total_gb * 100) if total_gb > 0 else 0
        ws_summary[f'A{row_num}'] = dir_name
        ws_summary[f'B{row_num}'] = round(data['Size GB'], 2)
        ws_summary[f'C{row_num}'] = f"{round(pct, 1)}%"
        row_num += 1

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 12

    wb.save(output_path)

    print("\n" + "="*80)
    print("TOP 20 DIRECTORIES BY SIZE:")
    print("="*80)
    for i, (dir_name, data) in enumerate(sorted_dirs[:20], 1):
        pct = (data['Size GB'] / total_gb * 100) if total_gb > 0 else 0
        print(f"{i:2}. {dir_name:30} {data['Size GB']:>8.2f} GB ({pct:>5.1f}%) - {data['Files']:>8,} files")

    print("\n" + "="*80)
    print(f"TOTAL SPACE ANALYZED: {total_gb:.2f} GB")
    print(f"TOTAL FILES FOUND: {total_files:,}")
    print(f"ACCESS ERRORS: {errors}")
    print("="*80)
    print(f"\nDetailed report saved: {output_path}")


# ---------------------------------------------------------------------------
# TASK: Scan C drive and list all high size files
# ---------------------------------------------------------------------------

def scan_c_drive_large_files(output_path: Path = None, min_size_mb: float = 100.0, max_results: int = 1000):
    """
    Recursively scans C: drive and lists all large files.
    
    Args:
        output_path: Where to save the report. Defaults to OUTPUT_DIR/large_files_c_drive.xlsx
        min_size_mb: Minimum file size in MB to report (default: 100 MB)
        max_results: Maximum number of files to report (default: 1000)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl required. Run: pip install openpyxl")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "large_files_c_drive.xlsx"

    output_path = Path(output_path)

    print(f"Scanning C: drive for files > {min_size_mb}MB...")
    print("This may take several minutes depending on drive size and file count...")

    files = []
    skipped_dirs = []
    min_size_bytes = min_size_mb * (1024 * 1024)
    c_drive = Path("C:\\")
    
    # Common directories to skip (speeds up scan)
    skip_patterns = [
        r'$Recycle.Bin', 'System Volume Information', 'ProgramData', 
        'hiberfil.sys', 'pagefile.sys', 'swapfile.sys',
        'Windows\\System32', 'Windows\\WinSxS', '.git', '__pycache__'
    ]

    def should_skip(path_str):
        for pattern in skip_patterns:
            if pattern.lower() in path_str.lower():
                return True
        return False

    try:
        for root, dirs, filenames in os.walk(c_drive, topdown=True):
            # Filter out directories to skip
            dirs[:] = [d for d in dirs if not should_skip(os.path.join(root, d))]
            
            if should_skip(root):
                skipped_dirs.append(root)
                continue
            
            for filename in filenames:
                if len(files) >= max_results:
                    print(f"Reached max results limit ({max_results})")
                    break
                
                try:
                    filepath = Path(root) / filename
                    file_size = filepath.stat().st_size
                    
                    if file_size >= min_size_bytes:
                        size_mb = file_size / (1024 * 1024)
                        size_gb = file_size / (1024 * 1024 * 1024)
                        file_ext = filepath.suffix if filepath.suffix else 'No Extension'
                        
                        files.append({
                            'File Path': str(filepath),
                            'File Name': filename,
                            'Size MB': round(size_mb, 2),
                            'Size GB': round(size_gb, 3),
                            'File Type': file_ext,
                            'Modified': filepath.stat().st_mtime,
                        })
                        
                        if len(files) % 100 == 0:
                            print(f"  Found {len(files)} files so far...")
                except (PermissionError, OSError):
                    pass
            
            if len(files) >= max_results:
                break

    except Exception as e:
        print(f"Error during scan: {e}")

    # Sort by size descending
    files.sort(key=lambda x: x['Size MB'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large Files"

    total_size_gb = sum(f['Size GB'] for f in files)
    ws['A1'] = f"C: Drive Large Files Report - {len(files)} files found ({total_size_gb:.2f} GB)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Files larger than {min_size_mb}MB | Scanned: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)

    headers = ['File Path', 'File Name', 'Size MB', 'Size GB', 'File Type', 'Modified']
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Color code by file size
    warning_fill = PatternFill("solid", fgColor="FFE699")  # Yellow for >1GB
    critical_fill = PatternFill("solid", fgColor="FF6B6B")  # Red for >5GB
    alt_fill = PatternFill("solid", fgColor="D9E1F2")  # Light blue

    for r_idx, file_data in enumerate(files, start=4):
        size_gb = file_data['Size GB']
        
        if size_gb > 5:
            fill = critical_fill
        elif size_gb > 1:
            fill = warning_fill
        else:
            fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        
        for col, key in enumerate(headers, start=1):
            value = file_data[key]
            # Format datetime
            if key == 'Modified':
                from datetime import datetime as dt
                value = dt.fromtimestamp(value).strftime('%Y-%m-%d %H:%M:%S')
            
            cell = ws.cell(row=r_idx, column=col, value=value)
            cell.fill = fill
            if key in ['Size MB', 'Size GB']:
                cell.alignment = Alignment(horizontal="right")
            elif key in ['File Type']:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

    col_widths = [50, 30, 12, 10, 15, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Large Files Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    ws_summary['A3'] = "Metric"
    ws_summary['B3'] = "Value"
    for col in ['A', 'B']:
        ws_summary[f'{col}3'].font = Font(bold=True)
        ws_summary[f'{col}3'].fill = PatternFill("solid", fgColor="1F4E79")
        ws_summary[f'{col}3'].font = Font(bold=True, color="FFFFFF")

    summary_data = [
        ('Total Files Found', len(files)),
        ('Total Size (GB)', round(total_size_gb, 2)),
        ('Average File Size (MB)', round(sum(f['Size MB'] for f in files) / len(files), 2) if files else 0),
        ('Largest File (GB)', round(files[0]['Size GB'], 3) if files else 0),
        ('Largest File Name', files[0]['File Name'] if files else 'N/A'),
    ]
    
    row_num = 4
    for metric, value in summary_data:
        ws_summary[f'A{row_num}'] = metric
        ws_summary[f'B{row_num}'] = value
        row_num += 1
    
    # File type distribution
    ws_summary['A11'] = "File Type Distribution"
    ws_summary['A11'].font = Font(bold=True, size=12)
    
    ws_summary['A13'] = "File Type"
    ws_summary['B13'] = "Count"
    ws_summary['C13'] = "Total Size GB"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}13'].font = Font(bold=True)
        ws_summary[f'{col}13'].fill = PatternFill("solid", fgColor="1F4E79")
        ws_summary[f'{col}13'].font = Font(bold=True, color="FFFFFF")
    
    file_types = {}
    for f in files:
        ftype = f['File Type']
        if ftype not in file_types:
            file_types[ftype] = {'count': 0, 'size': 0}
        file_types[ftype]['count'] += 1
        file_types[ftype]['size'] += f['Size GB']
    
    row_num = 14
    for ftype in sorted(file_types.keys(), key=lambda x: file_types[x]['size'], reverse=True):
        ws_summary[f'A{row_num}'] = ftype
        ws_summary[f'B{row_num}'] = file_types[ftype]['count']
        ws_summary[f'C{row_num}'] = round(file_types[ftype]['size'], 2)
        row_num += 1

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15

    wb.save(output_path)
    print(f"\n[SUCCESS] Large files scan complete!")
    print(f"  Files found: {len(files)}")
    print(f"  Total size: {total_size_gb:.2f} GB")
    print(f"  Output: {output_path}")


# ---------------------------------------------------------------------------
# TASK: Move .exe files from C: to D: drive
# ---------------------------------------------------------------------------

def move_exe_files_to_d_drive(dry_run: bool = True, exclude_system: bool = True, min_size_mb: float = 0):
    """
    Moves .exe files from C: to D: drive.
    
    Args:
        dry_run: If True, only shows what would be moved (don't actually move)
        exclude_system: If True, skips system directories (Windows, Program Files, etc.)
        min_size_mb: Only move .exe files larger than this size (0 = all files)
    """
    try:
        import shutil
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl required. Run: pip install openpyxl")
        return

    print("="*80)
    print(f"EXE FILE MOVE OPERATION - {'DRY RUN' if dry_run else 'ACTUAL MOVE'}")
    print("="*80)

    min_size_bytes = min_size_mb * (1024 * 1024)
    c_drive = Path("C:\\")
    d_drive = Path("D:\\")
    
    # System directories to exclude
    exclude_dirs = [
        'Windows', 'Program Files', 'Program Files (x86)', 'System32',
        'ProgramData', 'AppData', 'System Volume Information', '$Recycle.Bin'
    ]

    print(f"\nScanning C: drive for .exe files...")
    print(f"  Exclude system dirs: {exclude_system}")
    print(f"  Minimum size: {min_size_mb}MB")
    print()

    exe_files = []
    errors = []

    try:
        for root, dirs, filenames in os.walk(c_drive, topdown=True):
            # Filter directories
            if exclude_system:
                dirs[:] = [d for d in dirs if d not in exclude_dirs]
            
            for filename in filenames:
                if filename.lower().endswith('.exe'):
                    try:
                        filepath = Path(root) / filename
                        file_size = filepath.stat().st_size
                        
                        if file_size >= min_size_bytes:
                            size_mb = file_size / (1024 * 1024)
                            
                            # Create destination path preserving structure
                            rel_path = filepath.relative_to(c_drive)
                            dest_path = d_drive / rel_path
                            
                            exe_files.append({
                                'Source': str(filepath),
                                'Destination': str(dest_path),
                                'Size MB': round(size_mb, 2),
                                'Status': 'Ready to move',
                            })
                    except (PermissionError, OSError) as e:
                        errors.append({
                            'File': str(Path(root) / filename),
                            'Error': str(e),
                        })
    except Exception as e:
        print(f"Error during scan: {e}")

    # Sort by size
    exe_files.sort(key=lambda x: x['Size MB'], reverse=True)

    print(f"Found {len(exe_files)} .exe files to move")
    print(f"Total size: {sum(f['Size MB'] for f in exe_files):.2f} MB")
    
    if errors:
        print(f"Errors encountered: {len(errors)}")

    # Show preview
    print("\n" + "-"*80)
    print("TOP 20 FILES TO MOVE:")
    print("-"*80)
    for i, exe in enumerate(exe_files[:20], 1):
        print(f"{i:2}. {exe['Source'][-60:]:60} ({exe['Size MB']:>8.2f} MB)")

    if len(exe_files) > 20:
        print(f"    ... and {len(exe_files) - 20} more files")

    if dry_run:
        print("\n" + "="*80)
        print("DRY RUN COMPLETE - No files moved")
        print("="*80)
        print("\nTo proceed with actual move, call:")
        print(f"  move_exe_files_to_d_drive(dry_run=False, exclude_system=True, min_size_mb=0)")
        
        # Create a report anyway
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        report_path = OUTPUT_DIR / "exe_move_dryrun_report.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DRY RUN Report"

        ws['A1'] = "EXE Files Move - DRY RUN REPORT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Total files: {len(exe_files)} | Total size: {sum(f['Size MB'] for f in exe_files):.2f} MB"
        ws['A2'].font = Font(italic=True)

        headers = ['Source Path', 'Destination Path', 'Size MB', 'Status']
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        alt_fill = PatternFill("solid", fgColor="D9E1F2")
        for r_idx, exe in enumerate(exe_files, start=4):
            fill = alt_fill if r_idx % 2 == 0 else PatternFill()
            for col, key in enumerate(['Source', 'Destination', 'Size MB', 'Status'], start=1):
                cell = ws.cell(row=r_idx, column=col, value=exe.get(key))
                cell.fill = fill
                if key == 'Size MB':
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        col_widths = [50, 50, 12, 15]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w

        wb.save(report_path)
        print(f"\nDry-run report saved: {report_path}")
        
    else:
        # Actual move
        print("\n" + "="*80)
        print("MOVING FILES...")
        print("="*80)

        moved = []
        failed = []

        for i, exe in enumerate(exe_files, 1):
            src = Path(exe['Source'])
            dest = Path(exe['Destination'])
            
            try:
                # Create destination directory
                dest.parent.mkdir(parents=True, exist_ok=True)
                
                # Move file
                shutil.move(str(src), str(dest))
                
                exe['Status'] = 'Moved'
                moved.append(exe)
                
                if i % 10 == 0 or i == len(exe_files):
                    print(f"  [{i}/{len(exe_files)}] Moved {exe['Size MB']:.2f} MB")
                    
            except Exception as e:
                exe['Status'] = f'Failed: {str(e)}'
                failed.append(exe)
                print(f"  ERROR: {src.name} - {e}")

        print("\n" + "="*80)
        print(f"MOVE COMPLETE")
        print("="*80)
        print(f"Successfully moved: {len(moved)}")
        print(f"Failed: {len(failed)}")
        print(f"Total size moved: {sum(e['Size MB'] for e in moved):.2f} MB")

        # Create detailed report
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        report_path = OUTPUT_DIR / f"exe_move_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # Moved sheet
        ws_moved = wb.active
        ws_moved.title = "Moved"
        ws_moved['A1'] = f"Successfully Moved: {len(moved)} files"
        ws_moved['A1'].font = Font(bold=True, size=14, color="008000")

        headers = ['Source Path', 'Destination Path', 'Size MB', 'Status']
        header_fill = PatternFill("solid", fgColor="70AD47")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, h in enumerate(headers, start=1):
            cell = ws_moved.cell(row=2, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        for r_idx, exe in enumerate(moved, start=3):
            for col, key in enumerate(['Source', 'Destination', 'Size MB', 'Status'], start=1):
                cell = ws_moved.cell(row=r_idx, column=col, value=exe.get(key))
                if key == 'Size MB':
                    cell.alignment = Alignment(horizontal="right")

        # Failed sheet
        if failed:
            ws_failed = wb.create_sheet("Failed")
            ws_failed['A1'] = f"Failed: {len(failed)} files"
            ws_failed['A1'].font = Font(bold=True, size=14, color="FF0000")

            for col, h in enumerate(headers, start=1):
                cell = ws_failed.cell(row=2, column=col, value=h)
                cell.fill = PatternFill("solid", fgColor="FF6B6B")
                cell.font = header_font

            for r_idx, exe in enumerate(failed, start=3):
                for col, key in enumerate(['Source', 'Destination', 'Size MB', 'Status'], start=1):
                    cell = ws_failed.cell(row=r_idx, column=col, value=exe.get(key))
                    if key == 'Size MB':
                        cell.alignment = Alignment(horizontal="right")

        col_widths = [50, 50, 12, 20]
        for ws in [ws_moved] + ([ws_failed] if failed else []):
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[chr(64+i)].width = w

        wb.save(report_path)
        print(f"\nMove report saved: {report_path}")


# ---------------------------------------------------------------------------
# TASK: Export OneDrive .txt file contents to Excel
# ---------------------------------------------------------------------------

def export_onedrive_txt_contents_to_excel(onedrive_path: Path = None, output_path: Path = None):
    """
    Scans OneDrive for .txt files, exports full text to Excel, and groups similar text
    into separate tabs.

    Notes:
        - Excel cells are limited to 32,767 characters.
        - Text longer than that is split across multiple rows using Content Part.
    """
    try:
        import re
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("openpyxl required. Run: pip install openpyxl")
        return

    if onedrive_path is None:
        env_onedrive = os.environ.get("OneDrive")
        if env_onedrive:
            onedrive_path = Path(env_onedrive)
        else:
            onedrive_path = Path.home() / "OneDrive"

    onedrive_path = Path(onedrive_path)
    if not onedrive_path.exists():
        print(f"ERROR: OneDrive path not found: {onedrive_path}")
        return

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / "onedrive_notepad_text_export.xlsx"
    output_path = Path(output_path)

    print("=" * 80)
    print("ONEDRIVE NOTEPAD TEXT EXPORT")
    print("=" * 80)
    print(f"Source OneDrive: {onedrive_path}")
    print(f"Output Excel:    {output_path}")
    print("File filter:     *.txt")
    print()

    txt_files = [p for p in onedrive_path.rglob("*.txt") if p.is_file()]
    total_files = len(txt_files)
    print(f"Found {total_files} .txt files")

    if total_files == 0:
        print("No .txt files found. Nothing to export.")
        return

    max_cell_chars = 32000
    processed = 0
    failed = 0
    last_status_time = time.time()
    process_start = last_status_time
    encodings_to_try = ("utf-8", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "latin-1")

    file_records = []

    for txt_file in txt_files:
        processed += 1
        now = time.time()
        if now - last_status_time >= 120:
            elapsed = now - process_start
            rate = processed / elapsed if elapsed > 0 else 0
            remaining = total_files - processed
            eta_seconds = int(remaining / rate) if rate > 0 else -1
            if eta_seconds >= 0:
                eta_text = f"{eta_seconds // 60}m {eta_seconds % 60}s"
            else:
                eta_text = "unknown"
            print(
                f"[STATUS] {processed}/{total_files} files processed | "
                f"Current: {txt_file} | ETA: {eta_text}"
            )
            last_status_time = now

        stat = txt_file.stat()
        modified_time = datetime.fromtimestamp(stat.st_mtime).isoformat(sep=" ")
        size_kb = round(stat.st_size / 1024, 2)

        text_content = None
        for encoding in encodings_to_try:
            try:
                text_content = txt_file.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
            except OSError:
                text_content = None
                break

        if text_content is None:
            failed += 1
            file_records.append(
                {
                    "path": str(txt_file),
                    "modified_time": modified_time,
                    "size_kb": size_kb,
                    "content": "[ERROR] Could not read file content with supported encodings.",
                    "read_error": True,
                }
            )
            continue

        file_records.append(
            {
                "path": str(txt_file),
                "modified_time": modified_time,
                "size_kb": size_kb,
                "content": text_content,
                "read_error": False,
            }
        )

    workbook = openpyxl.Workbook()
    headers = ["File Path", "Modified Time", "Size (KB)", "Content Part", "Text Content"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    standard_column_width = 25
    standard_row_height = 15

    def setup_sheet(worksheet):
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_letter in ["A", "B", "C", "D", "E"]:
            worksheet.column_dimensions[col_letter].width = standard_column_width
        worksheet.sheet_format.defaultRowHeight = standard_row_height
        worksheet.row_dimensions[1].height = standard_row_height
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.zoomScale = 90

    def write_records(worksheet, records):
        row = 2
        for record in records:
            text_content = record["content"] if record["content"] is not None else ""
            parts = [text_content[i:i + max_cell_chars] for i in range(0, len(text_content), max_cell_chars)]
            if not parts:
                parts = [""]
            for part_index, part_text in enumerate(parts, start=1):
                worksheet.cell(row=row, column=1, value=record["path"])
                worksheet.cell(row=row, column=2, value=record["modified_time"])
                worksheet.cell(row=row, column=3, value=record["size_kb"])
                worksheet.cell(row=row, column=4, value=part_index)
                compact_text = part_text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
                content_cell = worksheet.cell(row=row, column=5, value=compact_text)
                content_cell.alignment = Alignment(wrap_text=False, shrink_to_fit=True, vertical="center")
                worksheet.row_dimensions[row].height = standard_row_height
                row += 1

    category_keywords = {
        "Tasks_Todo": [
            "todo", "to do", "task", "checklist", "pending", "deadline", "follow up",
            "complete", "action item", "reminder",
        ],
        "Office_Documents": [
            "document", "doc", "draft", "letter", "memo", "reference", "summary",
            "template", "official", "print",
        ],
        "Technical_Code": [
            "python", "javascript", "typescript", "java", "c++", "api", "json", "sql",
            "github", "git", "script", "function", "code", "bug", "error", "stack trace",
            "server", "deploy", "build", "debug", "traceback", "terminal", "cmd", "powershell",
            "docker", "kubernetes", "pipeline", "ci", "cd",
        ],
        "IT_Infra_Network": [
            "ip", "dns", "router", "switch", "firewall", "network", "lan", "wan", "wifi",
            "port", "tcp", "udp", "vpn", "ssl", "tls", "hostname", "subnet",
        ],
        "Passwords_Accounts": [
            "password", "passcode", "username", "login", "signin", "otp", "pin", "credential",
            "account", "2fa", "authenticator", "security answer",
        ],
        "Finance": [
            "invoice", "payment", "bank", "account", "upi", "amount", "balance", "salary",
            "tax", "gst", "receipt", "bill", "transaction", "credit", "debit", "loan",
            "emi", "interest", "budget", "expense", "profit", "loss",
        ],
        "Legal_Compliance": [
            "agreement", "contract", "legal", "law", "clause", "policy", "compliance",
            "terms", "condition", "nda", "notice",
        ],
        "Contacts": [
            "phone", "mobile", "email", "address", "contact", "whatsapp", "telegram",
            "call", "name:", "alternate", "landline", "fax", "contact person",
        ],
        "HR_Recruitment": [
            "resume", "cv", "candidate", "interview", "hiring", "employee", "joining",
            "offer letter", "hr", "attendance", "leave", "payroll",
        ],
        "Web_Links": [
            "http://", "https://", "www.", ".com", ".org", ".net", "link", "url",
            ".in", ".io", ".co", ".gov", ".edu",
        ],
        "Meetings_Work": [
            "meeting", "agenda", "minutes", "discussion", "project", "client", "team",
            "status", "update", "target", "milestone", "owner", "stakeholder",
        ],
        "Shopping": [
            "buy", "shopping", "order", "cart", "price", "item", "qty", "quantity",
            "amazon", "flipkart", "grocery", "wishlist", "discount", "offer", "delivery",
        ],
        "Personal_Notes": [
            "note", "diary", "journal", "thought", "personal", "idea", "plan", "reflection",
            "goal", "habit",
        ],
        "Education": [
            "study", "exam", "lesson", "course", "class", "chapter", "homework", "assignment",
            "syllabus", "teacher", "student", "lecture", "tutorial", "quiz",
        ],
        "Health_Medical": [
            "doctor", "hospital", "medicine", "tablet", "dose", "symptom", "diagnosis",
            "treatment", "appointment", "report", "bp", "sugar",
        ],
        "Travel_Transport": [
            "travel", "trip", "train", "flight", "bus", "cab", "taxi", "booking",
            "ticket", "pnr", "hotel", "itinerary", "passport", "visa",
        ],
        "Media_Content": [
            "video", "photo", "image", "audio", "music", "youtube", "thumbnail", "editing",
            "reel", "shorts", "podcast",
        ],
        "Banking_Reference": [
            "ifsc", "swift", "branch", "account number", "beneficiary", "cheque", "neft",
            "rtgs", "imps",
        ],
    }

    def categorize_record(record):
        if record["read_error"]:
            return "Read_Errors"

        text_content = record["content"] or ""
        if not text_content.strip():
            return "Empty"

        lowered = text_content.lower()
        filename = Path(record["path"]).name.lower()
        best_category = "Uncategorized"
        best_score = 0

        for category, keywords in category_keywords.items():
            score = 0
            for keyword in keywords:
                score += lowered.count(keyword)
                score += filename.count(keyword)
            if score > best_score:
                best_score = score
                best_category = category

        return best_category if best_score > 0 else "Uncategorized"

    def safe_sheet_name(name):
        cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
        return cleaned[:31] if len(cleaned) > 31 else cleaned

    all_sheet = workbook.active
    all_sheet.title = "All TXT"
    setup_sheet(all_sheet)
    write_records(all_sheet, file_records)

    categorized_records = {}
    for record in file_records:
        category = categorize_record(record)
        if category not in categorized_records:
            categorized_records[category] = []
        categorized_records[category].append(record)

    for category_name in sorted(categorized_records.keys()):
        sheet = workbook.create_sheet(safe_sheet_name(category_name))
        setup_sheet(sheet)
        write_records(sheet, categorized_records[category_name])

    workbook.save(output_path)

    print()
    print("[SUCCESS] OneDrive notepad text export completed.")
    print(f"Processed files: {processed}")
    print(f"Failed files:    {failed}")
    print(f"Categories:      {len(categorized_records)}")
    for category_name in sorted(categorized_records.keys()):
        print(f"  - {category_name}: {len(categorized_records[category_name])} file(s)")
    print(f"Saved to:        {output_path}")


# =============================================================================
# CLEAN QUICK ACCESS
# =============================================================================

def clean_quick_access(keep_top=3, dry_run=False):
    """
    Cleans Windows Explorer Quick Access by:
      1. Reading actual usage counts from the Windows JumpList AutomaticDestinations file.
      2. Identifying the top-N most-accessed folders currently pinned/frequent in Quick Access.
      3. Keeping those top-N, removing/unpinning all others.
      4. Clearing recent files from Quick Access (only the file entries, not folders).

    Args:
        keep_top : Number of most-used folder items to retain (default 3).
        dry_run  : If True, print what would happen without making any changes.
    """
    import subprocess
    import json
    import struct

    print(f"\n{'='*70}")
    print(f"CLEAN QUICK ACCESS  (keep_top={keep_top}, dry_run={dry_run})")
    print(f"{'='*70}")

    # -------------------------------------------------------------------------
    # Step 1: Read current Quick Access FOLDER items via PowerShell Shell API
    # -------------------------------------------------------------------------
    ps_get_folders = r"""
$ErrorActionPreference = 'SilentlyContinue'
$shell = New-Object -ComObject Shell.Application
$qa = $shell.Namespace("shell:::{679f85cb-0220-4080-b29b-5540cc05aab6}")
$result = @()
foreach ($item in $qa.Items()) {
    $isDir = Test-Path -LiteralPath $item.Path -PathType Container
    if ($isDir) {
        $verbs = @($item.Verbs() | ForEach-Object { $_.Name })
        $isPinned   = ($verbs -contains "Unpin from &Quick access")
        $removeVerb = if ($isPinned) { "Unpin from &Quick access" } else { "Remo&ve from Quick access" }
        $result += [PSCustomObject]@{
            Name       = $item.Name
            Path       = $item.Path
            IsPinned   = $isPinned
            RemoveVerb = $removeVerb
        }
    }
}
$result | ConvertTo-Json -Depth 2 -Compress
"""
    res = subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_get_folders],
        capture_output=True, text=True
    )
    raw = res.stdout.strip()
    if not raw:
        print("[ERROR] Could not retrieve Quick Access items. Aborting.")
        return

    qa_folders_raw = json.loads(raw)
    if isinstance(qa_folders_raw, dict):
        qa_folders_raw = [qa_folders_raw]

    qa_folders = [
        {
            "name":        item["Name"],
            "path":        item["Path"].rstrip("\\"),
            "is_pinned":   item["IsPinned"],
            "remove_verb": item["RemoveVerb"],
            "access_count": 0,
        }
        for item in qa_folders_raw
    ]

    print(f"\nFound {len(qa_folders)} folder(s) in Quick Access.")

    # -------------------------------------------------------------------------
    # Step 2: Parse the JumpList DestList for per-folder access counts
    # -------------------------------------------------------------------------
    jl_path = (
        Path(os.environ.get("APPDATA", ""))
        / r"Microsoft\Windows\Recent\AutomaticDestinations"
        / "f01b4d95cf55d32a.automaticDestinations-ms"
    )

    access_map = {}   # normalised_path -> access_count
    if jl_path.exists():
        try:
            import olefile
            with olefile.OleFileIO(str(jl_path)) as ole:
                dl = ole.openstream("DestList").read()

            # DestList binary layout (version 6, Windows 10/11):
            #   Header  : 32 bytes
            #   Entries : each = 134 bytes (fixed) + path_length*2 bytes (UTF-16LE)
            #     +88 WORD  : access_count
            #     +108 DWORD: pin_status (0xFFFFFFFF = pinned by user)
            #     +128 WORD : path_length (chars)
            #     +134 ...  : path (UTF-16LE)
            offset = 32
            while offset + 134 < len(dl):
                try:
                    path_len = struct.unpack_from("<H", dl, offset + 128)[0]
                    if path_len == 0 or path_len > 512:
                        break
                    path_bytes = dl[offset + 134: offset + 134 + path_len * 2]
                    raw_path = path_bytes.decode("utf-16-le", errors="replace").rstrip("\x00")
                    access_count = struct.unpack_from("<H", dl, offset + 88)[0]
                    normalised = raw_path.rstrip("\\").lower()
                    # Keep the highest count seen for each path
                    if normalised not in access_map or access_count > access_map[normalised]:
                        access_map[normalised] = access_count
                    offset += 134 + path_len * 2
                except Exception:
                    break
            print(f"Parsed {len(access_map)} path(s) from JumpList DestList.")
        except Exception as exc:
            print(f"[WARN] Could not parse JumpList ({exc}). Using positional order as fallback.")

    # -------------------------------------------------------------------------
    # Step 3: Attach access counts to Quick Access folder items
    # -------------------------------------------------------------------------
    for item in qa_folders:
        full_key = item["path"].lower()
        # JumpList stores paths without drive letter (e.g. \GitRepos\... not D:\GitRepos\...)
        # Strip drive letter for secondary lookup
        if len(full_key) >= 2 and full_key[1] == ":":
            no_drive_key = full_key[2:]   # e.g. \gitrepos\forensics-security\...
        else:
            no_drive_key = full_key
        item["access_count"] = access_map.get(full_key, 0) or access_map.get(no_drive_key, 0)

    # Sort by access count descending — purely by usage frequency
    qa_folders.sort(key=lambda x: x["access_count"], reverse=True)

    keep_items   = qa_folders[:keep_top]
    remove_items = qa_folders[keep_top:]

    # -------------------------------------------------------------------------
    # Step 4: Report plan
    # -------------------------------------------------------------------------
    print(f"\n{'-'*50}")
    print(f"  TOP {keep_top} TO KEEP:")
    for rank, item in enumerate(keep_items, 1):
        pinned_tag = "[PINNED]" if item["is_pinned"] else "[frequent]"
        print(f"    #{rank}  {item['name']:<35} count={item['access_count']}  {pinned_tag}")
        print(f"         {item['path']}")

    print(f"\n  {len(remove_items)} ITEM(S) TO REMOVE:")
    for item in remove_items:
        pinned_tag = "[PINNED]" if item["is_pinned"] else "[frequent]"
        print(f"    - {item['name']:<35} count={item['access_count']}  {pinned_tag}")
        print(f"      {item['path']}")

    if dry_run:
        print(f"\n[DRY RUN] No changes made.")
        return

    if not remove_items:
        print(f"\n[INFO] Nothing to remove. Quick Access already has {len(qa_folders)} item(s).")
        return

    # -------------------------------------------------------------------------
    # Step 5: Remove unwanted folders from Quick Access via Shell API
    # -------------------------------------------------------------------------
    print(f"\n{'-'*50}")
    print("  Removing items...")

    # Build inline PowerShell that embeds paths directly (avoids param-passing issues)
    keep_paths_ps = " ,".join(f'"{i["path"]}"' for i in keep_items)
    ps_remove = f"""
$ErrorActionPreference = 'SilentlyContinue'
$keepPaths = @({keep_paths_ps})
$shell  = New-Object -ComObject Shell.Application
$qa     = $shell.Namespace("shell:::{{{'{679f85cb-0220-4080-b29b-5540cc05aab6}'}}}") 
$removed = 0; $failed = 0
foreach ($item in @($qa.Items())) {{
    $isDir = Test-Path -LiteralPath $item.Path -PathType Container
    if (-not $isDir) {{ continue }}
    $norm = $item.Path.TrimEnd('\\')
    if ($keepPaths -contains $norm) {{ continue }}
    $verb = $item.Verbs() | Where-Object {{ $_.Name -eq 'Unpin from &Quick access' }} | Select-Object -First 1
    if (-not $verb) {{
        $verb = $item.Verbs() | Where-Object {{ $_.Name -eq 'Remo&ve from Quick access' }} | Select-Object -First 1
    }}
    if ($verb) {{
        $verb.DoIt()
        Start-Sleep -Milliseconds 200
        Write-Host "REMOVED: $($item.Name)"
        $removed++
    }} else {{
        Write-Host "NO_VERB: $($item.Name)"
        $failed++
    }}
}}
Write-Host "DONE removed=$removed failed=$failed"
"""
    res2 = subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_remove],
        capture_output=True, text=True
    )
    for line in res2.stdout.strip().splitlines():
        print(f"  {line}")

    # -------------------------------------------------------------------------
    # Step 6: Clear ALL recent file entries from Quick Access
    # -------------------------------------------------------------------------
    # Remove recent .lnk files
    recent_dir = Path(os.environ.get("APPDATA", "")) / r"Microsoft\Windows\Recent"
    cleared_lnk = 0
    if recent_dir.exists():
        for lnk in recent_dir.glob("*.lnk"):
            try:
                lnk.unlink()
                cleared_lnk += 1
            except Exception:
                pass

    # Nuke AutomaticDestinations and CustomDestinations jumplists
    # (these are the real source of recent file entries in Quick Access)
    cleared_jl = 0
    for subdir in ["AutomaticDestinations", "CustomDestinations"]:
        jl_dir = recent_dir / subdir
        if jl_dir.exists():
            for f in jl_dir.iterdir():
                try:
                    f.unlink()
                    cleared_jl += 1
                except Exception:
                    pass

    print(f"\n  Cleared {cleared_lnk} recent .lnk file(s) and {cleared_jl} jumplist file(s).")

    # -------------------------------------------------------------------------
    # Step 7: Re-pin the keeper folders (clearing jumplists unpins them)
    # -------------------------------------------------------------------------
    keep_paths_ps = " ,".join(f'"{i["path"]}"' for i in keep_items)
    ps_repin = f"""
$ErrorActionPreference = 'SilentlyContinue'
$shell = New-Object -ComObject Shell.Application
$paths = @({keep_paths_ps})
$pinned = 0
foreach ($p in $paths) {{
    $folder = $shell.Namespace($p)
    if ($folder) {{
        $item = $folder.Self
        $verb = $item.Verbs() | Where-Object {{ $_.Name -eq 'Pin to &Quick access' }} | Select-Object -First 1
        if ($verb) {{ $verb.DoIt(); $pinned++; Write-Host "PINNED: $p" }}
        else {{ Write-Host "ALREADY PINNED: $p" }}
        Start-Sleep -Milliseconds 300
    }}
}}
Write-Host "Re-pinned $pinned folder(s)"
"""
    res3 = subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_repin],
        capture_output=True, text=True
    )
    for line in res3.stdout.strip().splitlines():
        print(f"  {line}")

    print(f"\n[DONE] Quick Access cleaned. Kept top {keep_top} folder(s), removed {len(remove_items)}, cleared {cleared_lnk} lnk + {cleared_jl} jumplist file(s).")
    print(f"{'='*70}\n")


# =============================================================================
# COMPARE LOCAL REPO WITH REMOTE
# =============================================================================

def compare_local_with_remote(
    remote_url="https://github.com/padmaimpex1-pixel/commonLaptop",
    local_path=None,
    branch="main",
    output_report=True,
):
    """
    Compares a local Git repository folder with a remote GitHub repo and
    produces a detailed diff report showing:
      - Files only in local  (not pushed / new)
      - Files only in remote (deleted locally / not pulled)
      - Files in both but with different content (modified)
      - Files identical in both

    Args:
        remote_url   : GitHub HTTPS URL of the remote repo.
        local_path   : Path to local repo root. If None, searches D:\\GitRepos
                       for a repo whose origin matches remote_url.
                       If still not found, offers to clone it.
        branch       : Branch to compare against (default "main").
        output_report: If True, saves a text report to D:\\Generated-Outputs.

    Returns:
        dict with keys: only_local, only_remote, modified, identical, report_path
    """
    import subprocess
    import hashlib
    import tempfile

    REMOTE_URL = remote_url.rstrip("/")
    repo_name  = REMOTE_URL.rstrip("/").split("/")[-1]

    print(f"\n{'='*70}")
    print(f"COMPARE LOCAL vs REMOTE: {repo_name}")
    print(f"  Remote : {REMOTE_URL}")
    print(f"  Branch : {branch}")
    print(f"{'='*70}")

    # -------------------------------------------------------------------------
    # Step 1: Find local repo
    # -------------------------------------------------------------------------
    def _find_local_repo(search_root: Path, url: str) -> Path | None:
        """Walk D:\\GitRepos looking for a git repo whose origin matches url."""
        for git_dir in search_root.rglob(".git"):
            if not git_dir.is_dir():
                continue
            candidate = git_dir.parent
            try:
                result = subprocess.run(
                    ["git", "-C", str(candidate), "remote", "get-url", "origin"],
                    capture_output=True, text=True, timeout=5
                )
                remote = result.stdout.strip().rstrip("/")
                if remote.lower() == url.lower():
                    return candidate
            except Exception:
                continue
        return None

    local_repo = None
    if local_path:
        local_repo = Path(local_path)
        if not (local_repo / ".git").exists():
            print(f"[ERROR] {local_path} is not a git repository.")
            return {}
    else:
        print("  Searching D:\\GitRepos for a local clone...")
        local_repo = _find_local_repo(Path(r"D:\GitRepos"), REMOTE_URL)

    if local_repo is None:
        print(f"  No local clone found. Cloning {REMOTE_URL} ...")
        clone_dest = Path(r"D:\GitRepos") / repo_name
        result = subprocess.run(
            ["git", "clone", "--depth", "1", "--branch", branch, REMOTE_URL, str(clone_dest)],
            capture_output=True, text=True
        )
        if result.returncode != 0:
            # Try without --branch in case default branch differs
            result = subprocess.run(
                ["git", "clone", "--depth", "1", REMOTE_URL, str(clone_dest)],
                capture_output=True, text=True
            )
        if result.returncode == 0:
            local_repo = clone_dest
            print(f"  Cloned to: {local_repo}")
        else:
            print(f"[ERROR] Clone failed:\n{result.stderr}")
            return {}

    print(f"  Local  : {local_repo}")

    # -------------------------------------------------------------------------
    # Step 2: Fetch latest remote state into a temp worktree
    # -------------------------------------------------------------------------
    print("  Fetching remote...")
    fetch = subprocess.run(
        ["git", "-C", str(local_repo), "fetch", "origin"],
        capture_output=True, text=True
    )
    if fetch.returncode != 0:
        print(f"[WARN] Fetch failed: {fetch.stderr.strip()}")

    # Determine the remote tracking branch
    remote_ref = f"origin/{branch}"
    check_ref = subprocess.run(
        ["git", "-C", str(local_repo), "rev-parse", "--verify", remote_ref],
        capture_output=True, text=True
    )
    if check_ref.returncode != 0:
        # Try to find any remote branch
        branches = subprocess.run(
            ["git", "-C", str(local_repo), "branch", "-r"],
            capture_output=True, text=True
        ).stdout.strip().splitlines()
        if branches:
            remote_ref = branches[0].strip()
            print(f"  Using remote branch: {remote_ref}")
        else:
            print("[ERROR] No remote branches found.")
            return {}

    # -------------------------------------------------------------------------
    # Step 3: Get file trees for LOCAL HEAD and REMOTE via git ls-tree
    # -------------------------------------------------------------------------
    def _get_file_tree(repo: Path, ref: str) -> dict:
        """Returns {relative_path: sha1_blob} for all files at given ref."""
        result = subprocess.run(
            ["git", "-C", str(repo), "ls-tree", "-r", "--full-tree", ref],
            capture_output=True, text=True
        )
        tree = {}
        for line in result.stdout.splitlines():
            # format: <mode> <type> <sha>\t<path>
            parts = line.split("\t", 1)
            if len(parts) == 2:
                meta, path = parts
                sha = meta.split()[2]
                tree[path] = sha
        return tree

    local_ref  = "HEAD"
    local_tree  = _get_file_tree(local_repo, local_ref)
    remote_tree = _get_file_tree(local_repo, remote_ref)

    if not local_tree and not remote_tree:
        print("[ERROR] Could not read file trees from repo.")
        return {}

    print(f"  Local  HEAD : {len(local_tree)} file(s)")
    print(f"  Remote {remote_ref}: {len(remote_tree)} file(s)")

    # -------------------------------------------------------------------------
    # Step 4: Classify files
    # -------------------------------------------------------------------------
    all_paths    = set(local_tree) | set(remote_tree)
    only_local   = []   # in local, not in remote
    only_remote  = []   # in remote, not in local
    modified     = []   # in both, different blob SHA
    identical    = []   # in both, same blob SHA

    for path in sorted(all_paths):
        in_local  = path in local_tree
        in_remote = path in remote_tree
        if in_local and not in_remote:
            only_local.append(path)
        elif in_remote and not in_local:
            only_remote.append(path)
        elif local_tree[path] == remote_tree[path]:
            identical.append(path)
        else:
            modified.append(path)

    # -------------------------------------------------------------------------
    # Step 5: Get actual diff stats for modified files
    # -------------------------------------------------------------------------
    diff_stats = {}
    if modified:
        diff_result = subprocess.run(
            ["git", "-C", str(local_repo), "diff", "--stat", remote_ref, local_ref, "--"],
            capture_output=True, text=True
        )
        diff_stats["summary"] = diff_result.stdout.strip()

    # -------------------------------------------------------------------------
    # Step 6: Print report
    # -------------------------------------------------------------------------
    print(f"\n{'='*70}")
    print(f"  SUMMARY")
    print(f"{'='*70}")
    print(f"  Identical  : {len(identical):>4} file(s)")
    print(f"  Modified   : {len(modified):>4} file(s)  (local differs from remote)")
    print(f"  Only LOCAL : {len(only_local):>4} file(s)  (not pushed to remote)")
    print(f"  Only REMOTE: {len(only_remote):>4} file(s)  (not pulled to local)")

    if only_local:
        print(f"\n  --- ONLY IN LOCAL (not pushed) ---")
        for f in only_local:
            print(f"    + {f}")

    if only_remote:
        print(f"\n  --- ONLY IN REMOTE (not in local) ---")
        for f in only_remote:
            print(f"    - {f}")

    if modified:
        print(f"\n  --- MODIFIED (content differs) ---")
        for f in modified:
            print(f"    ~ {f}")
        if diff_stats.get("summary"):
            print(f"\n  Diff stats:")
            for line in diff_stats["summary"].splitlines():
                print(f"    {line}")

    # -------------------------------------------------------------------------
    # Step 7: Save report to D:\Generated-Outputs
    # -------------------------------------------------------------------------
    report_path = None
    if output_report:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = OUTPUT_DIR / f"repo_compare_{repo_name}_{ts}.txt"
        lines = [
            f"REPO COMPARE REPORT",
            f"Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Remote    : {REMOTE_URL}",
            f"Local     : {local_repo}",
            f"Branch    : {branch}",
            f"",
            f"SUMMARY",
            f"  Identical   : {len(identical)}",
            f"  Modified    : {len(modified)}",
            f"  Only local  : {len(only_local)}",
            f"  Only remote : {len(only_remote)}",
            f"",
        ]
        if only_local:
            lines += ["ONLY IN LOCAL (not pushed):"]
            lines += [f"  + {f}" for f in only_local]
            lines.append("")
        if only_remote:
            lines += ["ONLY IN REMOTE (not in local):"]
            lines += [f"  - {f}" for f in only_remote]
            lines.append("")
        if modified:
            lines += ["MODIFIED (content differs):"]
            lines += [f"  ~ {f}" for f in modified]
            if diff_stats.get("summary"):
                lines += ["", "DIFF STATS:"]
                lines += [f"  {l}" for l in diff_stats["summary"].splitlines()]
            lines.append("")
        if identical:
            lines += ["IDENTICAL FILES:"]
            lines += [f"  = {f}" for f in identical]

        report_path.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n  Report saved: {report_path}")

    print(f"\n[DONE] Compare complete.")
    print(f"{'='*70}\n")

    return {
        "only_local":  only_local,
        "only_remote": only_remote,
        "modified":    modified,
        "identical":   identical,
        "report_path": report_path,
    }
